import openpyxl

book = openpyxl.load_workbook("/Users/aelya/Documents/PythonDemo.xlsx")
sheet = book.active
Dict = {}

cell = sheet.cell(row=1, column=2)
print(cell.value)
firstName = sheet.cell(row=2, column=2).value = "Aelya"
print(firstName)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value)

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)

# to print certain row
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column + 1): #to get columns
            #Dic["lastnmae" = "Askary
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)

