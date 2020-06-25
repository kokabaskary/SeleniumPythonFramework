import openpyxl


class HomePageData:

    #[("Aelya", "Askary", "Female"),("Hassan", "Askary", "Male")]) TUPLE DATA TYPE
    test_HomePage_data = [{"Name":"Aelya", "email":"Askary", "gender":"Female"},
                            {"Name":"Hassan", "email":"Askary", "gender":"Male"}]



    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("/Users/aelya/Documents/PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]